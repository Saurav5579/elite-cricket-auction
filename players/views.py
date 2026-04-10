from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.utils import timezone
from datetime import timedelta
from django.conf import settings
import razorpay

from .models import Player, Auction, Bid
from .forms import PlayerForm
import razorpay

from django.shortcuts import render, get_object_or_404
from .models import Player

# =========================
# HOME PAGE
# =========================
def home(request):
    return render(request, "players/home.html")


# =========================
# PLAYER REGISTER
# =========================
def register_player(request):

    if request.method == "POST":

        form = PlayerForm(request.POST, request.FILES)

        mobile = request.POST.get("mobile")
        email = request.POST.get("email")

        # =========================
        # 🔥 RESUME REGISTRATION (IMPORTANT)
        # =========================
        existing_player = Player.objects.filter(
            mobile=mobile,
            payment_status=False
        ).first()

        if existing_player:
            return redirect("payment_page", player_id=existing_player.id)

        # =========================
        # ✅ DUPLICATE CHECK (ONLY PAID USERS)
        # =========================
        if Player.objects.filter(mobile=mobile, payment_status=True).exists():
            return render(request, "players/register.html", {
                "form": form,
                "error": "This mobile number is already registered!"
            })

        if Player.objects.filter(email=email, payment_status=True).exists():
            return render(request, "players/register.html", {
                "form": form,
                "error": "This email is already registered!"
            })

        # =========================
        # ✅ SAVE NEW PLAYER
        # =========================
        if form.is_valid():

            player = form.save(commit=False)

            player.payment_status = False

            player.save()

            return redirect("payment_page", player_id=player.id)

    else:
        form = PlayerForm()

    return render(request, "players/register.html", {"form": form})

# =========================
# PAYMENT PAGE (Direct Razorpay Popup)
# =========================

import razorpay
from django.shortcuts import render, get_object_or_404
from django.conf import settings
from .models import Player


def payment_page(request, player_id):

    # ✅ Player fetch
    player = get_object_or_404(Player, id=player_id)

    # ✅ Razorpay client
    client = razorpay.Client(
        auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET)
    )

    # =========================
    # 🔥 USE SETTINGS (GLOBAL CONTROL)
    # =========================
    final_amount = settings.REGISTRATION_FEE

    # Razorpay paise me leta hai (₹ → paise)
    amount = int(final_amount * 100)

    # ✅ Create order
    order = client.order.create({
        "amount": amount,
        "currency": "INR",
        "payment_capture": "1"
    })

    # ✅ Context
    context = {
        "player": player,
        "razorpay_key": settings.RAZORPAY_KEY_ID,
        "amount": amount,
        "display_amount": final_amount,
        "order_id": order["id"]
    }

    return render(request, "players/payment.html", context)


# =========================
# PAYMENT SUCCESS (FINAL CLEAN VERSION)
# =========================

from django.shortcuts import get_object_or_404, render
from django.conf import settings
from django.utils import timezone
import razorpay
import threading
from .models import Player
from .utils import send_registration_emails


def payment_success(request, player_id):

    player = get_object_or_404(Player, id=player_id)

    razorpay_payment_id = request.POST.get("razorpay_payment_id") or request.GET.get("razorpay_payment_id")
    razorpay_order_id = request.POST.get("razorpay_order_id") or request.GET.get("razorpay_order_id")
    razorpay_signature = request.POST.get("razorpay_signature") or request.GET.get("razorpay_signature")

    error_message = None

    # ✅ Amount safe
    final_amount = player.payment_amount if player.payment_amount else settings.REGISTRATION_FEE

    try:
        # =========================
        # VERIFY PAYMENT
        # =========================
        if razorpay_order_id and razorpay_payment_id and razorpay_signature:

            client = razorpay.Client(
                auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET)
            )

            client.utility.verify_payment_signature({
                "razorpay_order_id": razorpay_order_id,
                "razorpay_payment_id": razorpay_payment_id,
                "razorpay_signature": razorpay_signature
            })

            print("✅ Payment verified")

        else:
            print("⚠️ Missing Razorpay params")

        # =========================
        # SAVE PAYMENT (ONLY ONCE)
        # =========================
        if not player.payment_status:

            player.payment_status = True
            player.payment_date = timezone.now()

            # ✅ SAFE Transaction ID
            if razorpay_payment_id:
                player.transaction_id = razorpay_payment_id
            elif razorpay_order_id:
                player.transaction_id = razorpay_order_id
            else:
                player.transaction_id = "SUCCESS"

            player.payment_amount = settings.REGISTRATION_FEE
            player.save()

            print("✅ Payment saved")

            # =========================
            # SEND EMAIL (BACKGROUND)
            # =========================
            try:
                threading.Thread(
                    target=send_registration_emails,
                    args=(player,)
                ).start()
            except Exception as e:
                print("Email failed:", e)

    except Exception as e:
        print("❌ Payment verification error:", e)
        error_message = "Payment verification failed. Please contact support."

    # =========================
    # SUCCESS PAGE
    # =========================
    return render(request, "players/payment_success.html", {
        "player": player,
        "amount": player.payment_amount,
        "transaction_id": player.transaction_id,
        "error": error_message
    })
# =========================
# PLAYER LIST
# =========================
from .models import SiteSetting

def player_list(request):
    players = Player.objects.all().order_by("player_id")
    settings_obj = SiteSetting.objects.first()

    return render(request, "players/player_list.html", {
        "players": players,
        "settings_obj": settings_obj
    })


# =========================
# START AUCTION
# =========================
def start_auction(request, player_id):

    if not getattr(settings, "AUCTION_ENABLED", False):
        return redirect("player_list")

    player = get_object_or_404(Player, id=player_id)

    if not player.payment_status:
        return redirect("payment_page", player_id=player.id)

    active = Auction.objects.filter(is_active=True).first()

    if active:
        if timezone.now() > active.end_time:
            active.is_active = False
            active.is_closed = True
            active.save()
        else:
            return redirect("live_auction", auction_id=active.id)

    auction = Auction.objects.create(
        player=player,
        current_price=player.base_price,
        start_time=timezone.now(),
        end_time=timezone.now() + timedelta(seconds=60),
        is_active=True,
        is_closed=False
    )

    return redirect("live_auction", auction_id=auction.id)


# =========================
# LIVE AUCTION PAGE
# =========================
def live_auction(request, auction_id):

    auction = get_object_or_404(Auction, id=auction_id)

    if auction.is_active and timezone.now() > auction.end_time:

        auction.is_active = False
        auction.is_closed = True

        if auction.highest_bidder:
            auction.player.is_sold = True
            auction.player.sold_price = auction.current_price
            auction.player.save()

        auction.save()

    return render(request, "auction/live.html", {"auction": auction})


# =========================
# PLACE BID (AJAX)
# =========================
def place_bid(request, auction_id):

    if request.method != "POST":
        return JsonResponse({"status": "error"})

    auction = get_object_or_404(Auction, id=auction_id)

    if not auction.is_active:
        return JsonResponse({"status": "closed"})

    try:
        amount = int(request.POST.get("amount", 0))
    except ValueError:
        return JsonResponse({"status": "bad_amount"})

    bidder = request.POST.get("bidder", "Guest")

    if amount > auction.current_price:

        auction.current_price = amount
        auction.highest_bidder = bidder
        auction.save()

        Bid.objects.create(
            auction=auction,
            bidder=bidder,
            amount=amount
        )

        return JsonResponse({
            "status": "ok",
            "price": amount,
            "bidder": bidder
        })

    return JsonResponse({"status": "low"})


# =========================
# REALTIME STATE API
# =========================
def auction_state(request, auction_id):

    auction = get_object_or_404(Auction, id=auction_id)

    bids = Bid.objects.filter(auction=auction).order_by("-time")[:5]

    bid_list = [
        {
            "bidder": b.bidder,
            "amount": b.amount
        }
        for b in bids
    ]

    return JsonResponse({
        "price": auction.current_price,
        "bidder": auction.highest_bidder,
        "active": auction.is_active,
        "bids": bid_list
    })


# =========================
# EXPORT PLAYERS TO EXCEL
# =========================

import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill, Alignment
from django.contrib.auth.decorators import login_required
from .models import Player


@login_required
def export_players_excel(request):

    if not request.user.is_superuser:
        return HttpResponse("Access Denied")

    players = Player.objects.all().order_by("player_id")

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Registered Players"

    headers = [
        "Player ID",
        "Name",
        "Age",
        "City",
        "Mobile",
        "Email",
        "Role",
        "Playing Style",
        "Experience",
        "Base Price",
        "Payment Status",
        "Transaction ID",
        "Payment Date"
    ]

    sheet.append(headers)

    # Header style
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Status colors
    paid_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    pending_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    row_num = 2

    for player in players:

        payment_status = "Paid" if player.payment_status else "Pending"

        sheet.append([
            player.player_id,
            player.name,
            player.age,
            player.city,
            player.mobile,
            player.email,
            player.role,
            player.playing_style,
            player.experience,
            player.base_price,
            payment_status,
            player.transaction_id if hasattr(player, "transaction_id") and player.transaction_id else "",
            str(player.payment_date) if player.payment_date else ""
        ])

        status_cell = sheet.cell(row=row_num, column=11)

        if payment_status == "Paid":
            status_cell.fill = paid_fill
        else:
            status_cell.fill = pending_fill

        row_num += 1

    # Auto column width
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        sheet.column_dimensions[column_letter].width = max_length + 4

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = "attachment; filename=players.xlsx"

    workbook.save(response)

    return response

def terms(request):
    return render(request, "players/terms.html")



def player_detail(request, id):
    from .models import Player
    from django.shortcuts import render, get_object_or_404

    player = get_object_or_404(Player, id=id)

    return render(request, "players/player_detail.html", {
        "player": player
    })

from django.contrib.auth import get_user_model
from django.http import HttpResponse

def create_admin(request):
    User = get_user_model()

    if not User.objects.filter(username='admin').exists():
        User.objects.create_superuser(
            username='admin',
            email='admin@gmail.com',
            password='admin123'
        )
        return HttpResponse("Admin Created")

    return HttpResponse("Already Exists")
